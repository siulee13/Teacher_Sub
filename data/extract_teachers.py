import openpyxl
import sqlite3
import json

wb = openpyxl.load_workbook('/Users/siu_lee/Desktop/KFPS/KFPS/校務/25-26/教師人名表25-26.xlsx', data_only=True)

# Read from 教師名單 sheet (has English initials)
ws = wb['教師名單']
teachers = []

# Parse class teachers (columns A-D for 1A-3F, E-H for 4A-6F)
for row in ws.iter_rows(min_row=7, max_row=30, values_only=True):
    vals = [str(c).strip() if c else '' for c in row]
    # Left side: col 0=class, 1=fullname, 2=shortname, 3=initials
    if vals[0] and vals[1]:
        cls = vals[0].replace('(', '(').replace(')', ')').replace(' ', '')
        name = vals[1].replace(' ', '').replace('\u3000', '')
        short = vals[2].replace(' ', '') if vals[2] else ''
        initials = vals[3].strip() if vals[3] else ''
        role = '副班主任' if '(副)' in cls or '(副)' in cls else '班主任'
        home = cls.replace('(正)', '').replace('(副)', '').replace('(正)', '').replace('(副)', '')
        teachers.append({'fullName': name, 'shortName': short, 'initials': initials, 'role': role, 'homeClass': home})
    # Right side: col 4=class, 5=fullname, 6=shortname, 7=initials
    if vals[4] and vals[5]:
        cls = vals[4].replace('(', '(').replace(')', ')').replace(' ', '')
        name = vals[5].replace(' ', '').replace('\u3000', '')
        short = vals[6].replace(' ', '') if vals[6] else ''
        initials = vals[7].strip() if vals[7] else ''
        role = '副班主任' if '(副)' in cls or '(副)' in cls else '班主任'
        home = cls.replace('(正)', '').replace('(副)', '').replace('(正)', '').replace('(副)', '')
        teachers.append({'fullName': name, 'shortName': short, 'initials': initials, 'role': role, 'homeClass': home})

# Non-class teachers (columns 8-11)
for row in ws.iter_rows(min_row=7, max_row=15, values_only=True):
    vals = [str(c).strip() if c else '' for c in row]
    if len(vals) > 10 and vals[8] and vals[8] != 'None':
        name = vals[8].replace(' ', '').replace('\u3000', '')
        short = vals[9].replace(' ', '') if vals[9] else ''
        initials = vals[10].strip() if vals[10] else ''
        teachers.append({'fullName': name, 'shortName': short, 'initials': initials, 'role': '老師', 'homeClass': None})

# 副校長及主任 (columns 11-13)
for row in ws.iter_rows(min_row=7, max_row=20, values_only=True):
    vals = [str(c).strip() if c else '' for c in row]
    if len(vals) > 13 and vals[11] and vals[11] != 'None':
        name = vals[11].replace(' ', '').replace('\u3000', '')
        short = vals[12].replace(' ', '') if vals[12] else ''
        initials = vals[13].strip() if vals[13] else ''
        teachers.append({'fullName': name, 'shortName': short, 'initials': initials, 'role': '主任', 'homeClass': None})

# Deduplicate by fullName
seen = set()
unique = []
for t in teachers:
    if t['fullName'] not in seen and t['fullName']:
        seen.add(t['fullName'])
        unique.append(t)

# Create SQLite DB
conn = sqlite3.connect('/Users/siu_lee/Teacher_Sub/data/timetable.db')
c = conn.cursor()
c.execute('DROP TABLE IF EXISTS teacher_names')
c.execute('''CREATE TABLE teacher_names (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    fullName TEXT NOT NULL,
    shortName TEXT,
    initials TEXT,
    role TEXT,
    homeClass TEXT
)''')
for t in unique:
    c.execute('INSERT INTO teacher_names (fullName, shortName, initials, role, homeClass) VALUES (?, ?, ?, ?, ?)',
              (t['fullName'], t['shortName'], t['initials'], t['role'], t['homeClass']))
conn.commit()

print(f"Inserted {len(unique)} teachers:")
for t in unique:
    print(f"  {t['fullName']} ({t['shortName']}) [{t['initials']}] - {t['role']} {t['homeClass'] or ''}")

conn.close()
